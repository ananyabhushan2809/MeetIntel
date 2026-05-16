"""
export_db.py - Export Database to Excel
========================================
Reads all tables from the SQLite database and exports them
to a professional Excel file with formatted sheets.

Usage:
    python export_db.py

Output:
    Creates 'database_export.xlsx' in the backend folder
"""

import sqlite3
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Path to the database
DATABASE_PATH = os.path.join(os.path.dirname(__file__), 'meetings.db')
OUTPUT_FILE = os.path.join(os.path.dirname(__file__), 'database_export.xlsx')


def style_header(ws, headers, row=1):
    """Apply professional styling to header row."""
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def style_data_rows(ws, start_row, end_row, num_cols):
    """Apply alternating row colors and borders to data cells."""
    data_font = Font(name='Calibri', size=10)
    data_align = Alignment(vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )
    even_fill = PatternFill(start_color='F2F7FF', end_color='F2F7FF', fill_type='solid')

    for row_idx in range(start_row, end_row + 1):
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border
            # Alternate row coloring
            if (row_idx - start_row) % 2 == 1:
                cell.fill = even_fill


def auto_width(ws, num_cols, min_width=12, max_width=40):
    """Auto-adjust column widths based on content."""
    from openpyxl.utils import get_column_letter
    for col_idx in range(1, num_cols + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if cell.value:
                        cell_len = len(str(cell.value))
                        if cell_len > max_len:
                            max_len = cell_len
                except:
                    pass
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 4, max_width))


def export_database():
    """Export all database tables to a formatted Excel file."""

    if not os.path.exists(DATABASE_PATH):
        print("[ERROR] Database file not found! Run the app first to create it.")
        return

    conn = sqlite3.connect(DATABASE_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    wb = Workbook()

    # ===========================
    # Sheet 1: Users
    # ===========================
    ws_users = wb.active
    ws_users.title = "Users"
    ws_users.sheet_properties.tabColor = "2563EB"

    headers = ['ID', 'Full Name', 'Email Address', 'Registered']
    style_header(ws_users, headers)

    users = cursor.execute('SELECT * FROM users ORDER BY id').fetchall()
    for row_idx, user in enumerate(users, 2):
        ws_users.cell(row=row_idx, column=1, value=user['id'])
        ws_users.cell(row=row_idx, column=2, value=user['name'])
        ws_users.cell(row=row_idx, column=3, value=user['email'])
        ws_users.cell(row=row_idx, column=4, value='Yes')

    if users:
        style_data_rows(ws_users, 2, len(users) + 1, len(headers))
    auto_width(ws_users, len(headers))

    # Add summary row
    summary_row = len(users) + 3
    ws_users.cell(row=summary_row, column=1, value=f"Total Users: {len(users)}")
    ws_users.cell(row=summary_row, column=1).font = Font(name='Calibri', bold=True, size=11, color='2563EB')

    # ===========================
    # Sheet 2: Meetings
    # ===========================
    ws_meetings = wb.create_sheet("Meetings")
    ws_meetings.sheet_properties.tabColor = "10B981"

    headers = ['ID', 'Title', 'Summary', 'Created By (User ID)', 'Created At']
    style_header(ws_meetings, headers)

    meetings = cursor.execute('''
        SELECT m.*, u.name as user_name 
        FROM meetings m 
        LEFT JOIN users u ON m.created_by = u.id 
        ORDER BY m.id
    ''').fetchall()

    for row_idx, meeting in enumerate(meetings, 2):
        ws_meetings.cell(row=row_idx, column=1, value=meeting['id'])
        ws_meetings.cell(row=row_idx, column=2, value=meeting['title'])
        ws_meetings.cell(row=row_idx, column=3, value=meeting['summary'] or 'N/A')
        ws_meetings.cell(row=row_idx, column=4, value=f"{meeting['user_name']} (ID: {meeting['created_by']})")
        ws_meetings.cell(row=row_idx, column=5, value=meeting['created_at'])

    if meetings:
        style_data_rows(ws_meetings, 2, len(meetings) + 1, len(headers))
    auto_width(ws_meetings, len(headers))

    summary_row = len(meetings) + 3
    ws_meetings.cell(row=summary_row, column=1, value=f"Total Meetings: {len(meetings)}")
    ws_meetings.cell(row=summary_row, column=1).font = Font(name='Calibri', bold=True, size=11, color='10B981')

    # ===========================
    # Sheet 3: Tasks
    # ===========================
    ws_tasks = wb.create_sheet("Tasks")
    ws_tasks.sheet_properties.tabColor = "F59E0B"

    headers = ['ID', 'Task Name', 'Assigned To', 'Due Date', 'Status', 'Meeting ID', 'Created At']
    style_header(ws_tasks, headers)

    tasks = cursor.execute('SELECT * FROM tasks ORDER BY id').fetchall()
    for row_idx, task in enumerate(tasks, 2):
        ws_tasks.cell(row=row_idx, column=1, value=task['id'])
        ws_tasks.cell(row=row_idx, column=2, value=task['task_name'])
        ws_tasks.cell(row=row_idx, column=3, value=task['assigned_to'])
        ws_tasks.cell(row=row_idx, column=4, value=task['due_date'] or 'N/A')
        ws_tasks.cell(row=row_idx, column=5, value=task['status'])
        ws_tasks.cell(row=row_idx, column=6, value=task['meeting_id'] or 'N/A')
        ws_tasks.cell(row=row_idx, column=7, value=task['created_at'])

    # Color-code status cells
    status_colors = {
        'Pending': 'FEF3C7',
        'In Progress': 'DBEAFE',
        'Completed': 'D1FAE5',
    }
    for row_idx, task in enumerate(tasks, 2):
        status = task['status']
        if status in status_colors:
            ws_tasks.cell(row=row_idx, column=5).fill = PatternFill(
                start_color=status_colors[status],
                end_color=status_colors[status],
                fill_type='solid'
            )

    if tasks:
        style_data_rows(ws_tasks, 2, len(tasks) + 1, len(headers))
        # Re-apply status colors after general styling
        for row_idx, task in enumerate(tasks, 2):
            status = task['status']
            if status in status_colors:
                ws_tasks.cell(row=row_idx, column=5).fill = PatternFill(
                    start_color=status_colors[status],
                    end_color=status_colors[status],
                    fill_type='solid'
                )

    auto_width(ws_tasks, len(headers))

    summary_row = len(tasks) + 3
    ws_tasks.cell(row=summary_row, column=1, value=f"Total Tasks: {len(tasks)}")
    ws_tasks.cell(row=summary_row, column=1).font = Font(name='Calibri', bold=True, size=11, color='F59E0B')

    # ===========================
    # Sheet 4: Summary Dashboard
    # ===========================
    ws_summary = wb.create_sheet("Summary", 0)
    ws_summary.sheet_properties.tabColor = "4F46E5"

    # Title
    ws_summary.merge_cells('A1:D1')
    title_cell = ws_summary.cell(row=1, column=1, value="MeetIntel - Database Export Report")
    title_cell.font = Font(name='Calibri', bold=True, size=16, color='2563EB')
    title_cell.alignment = Alignment(horizontal='left', vertical='center')

    ws_summary.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    ws_summary.cell(row=2, column=1).font = Font(name='Calibri', size=10, color='64748B')

    # Stats
    stats_start = 4
    stat_headers = ['Metric', 'Count']
    style_header(ws_summary, stat_headers, row=stats_start)

    pending = len([t for t in tasks if t['status'] == 'Pending'])
    in_progress = len([t for t in tasks if t['status'] == 'In Progress'])
    completed = len([t for t in tasks if t['status'] == 'Completed'])

    stats = [
        ('Total Registered Users', len(users)),
        ('Total Meetings Uploaded', len(meetings)),
        ('Total Tasks Created', len(tasks)),
        ('', ''),
        ('Pending Tasks', pending),
        ('In Progress Tasks', in_progress),
        ('Completed Tasks', completed),
        ('', ''),
        ('Productivity Rate', f"{round((completed / len(tasks)) * 100, 1)}%" if tasks else '0%'),
    ]

    for idx, (metric, count) in enumerate(stats):
        row = stats_start + 1 + idx
        ws_summary.cell(row=row, column=1, value=metric)
        ws_summary.cell(row=row, column=2, value=count)
        if metric:
            ws_summary.cell(row=row, column=1).font = Font(name='Calibri', size=11)
            ws_summary.cell(row=row, column=2).font = Font(name='Calibri', size=11, bold=True)
            ws_summary.cell(row=row, column=2).alignment = Alignment(horizontal='center')

    auto_width(ws_summary, 2, min_width=25)

    # ===========================
    # Save
    # ===========================
    conn.close()
    wb.save(OUTPUT_FILE)

    print(f"[OK] Database exported successfully!")
    print(f"     File: {OUTPUT_FILE}")
    print(f"     Sheets: Summary, Users ({len(users)}), Meetings ({len(meetings)}), Tasks ({len(tasks)})")


if __name__ == '__main__':
    export_database()
